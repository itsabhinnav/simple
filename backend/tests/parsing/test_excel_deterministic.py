"""Tests for the deterministic Excel parser."""

from __future__ import annotations

import zipfile

from src.services.parsing.deterministic.excel_parser import ExcelDeterministicParser


def test_merged_cell_propagation(tmp_path, xlsx_fixture):
    parser = ExcelDeterministicParser(tmp_dir=str(tmp_path / "out"))
    result = parser.parse(xlsx_fixture)

    a4 = next((c for c in result.cells if c.ref == "A4"), None)
    b4 = next((c for c in result.cells if c.ref == "B4"), None)
    assert a4 is not None and a4.value == "Section Header"
    assert a4.is_merged_origin is True
    assert b4 is not None and b4.value == "Section Header"
    assert b4.inherited_from == "A4"


def test_dimension_clamping_prevents_million_row_scan(tmp_path, xlsx_fixture):
    parser = ExcelDeterministicParser(tmp_dir=str(tmp_path / "out"))
    result = parser.parse(xlsx_fixture)
    dim = result.dimensions.get("Spec", "")
    assert ":" in dim
    end_ref = dim.split(":", 1)[1]
    assert int("".join(ch for ch in end_ref if ch.isdigit())) < 1_000_000
    rows_scanned = {cell.row for cell in result.cells}
    assert max(rows_scanned) < 1_000_000


def test_hidden_columns_recorded_but_excluded(tmp_path, xlsx_fixture):
    parser = ExcelDeterministicParser(tmp_dir=str(tmp_path / "out"))
    result = parser.parse(xlsx_fixture)
    sheet_hidden = result.hidden.get("Spec", {})
    assert 4 in sheet_hidden.get("cols", []), f"col D (4) should be hidden, got {sheet_hidden}"
    for cell in result.cells:
        assert not (cell.col == 4 and cell.value is not None)


def test_image_anchor_resolved_to_a1_ref(tmp_path, xlsx_fixture):
    parser = ExcelDeterministicParser(tmp_dir=str(tmp_path / "out"))
    result = parser.parse(xlsx_fixture)
    assert len(result.images) == 1
    image = result.images[0]
    assert image.sheet == "Spec"
    assert image.anchor_cell == "C14"
    assert image.bounding_box_estimate is not None
    assert image.bounding_box_estimate.start_col == 3
    assert image.bounding_box_estimate.start_row == 14


def test_context_window_around_image(tmp_path, xlsx_fixture):
    parser = ExcelDeterministicParser(tmp_dir=str(tmp_path / "out"))
    result = parser.parse(xlsx_fixture)
    assert len(result.context_windows) == 1
    window = result.context_windows[0]
    assert window.sheet == "Spec"
    assert any("[IMAGE_" in cell for row in window.rows for cell in row)


def test_phantom_row_beyond_dimension_does_not_blow_up(tmp_path, png_factory):
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage

    img = png_factory("phantom.png", color=(10, 10, 10))
    xlsx_path = tmp_path / "phantom.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phantom"
    ws["A1"] = "header"
    ws["A2"] = "body"
    image = XLImage(str(img))
    ws.add_image(image, "B5")
    wb.save(str(xlsx_path))

    parser = ExcelDeterministicParser(tmp_dir=str(tmp_path / "out"))
    result = parser.parse(xlsx_path)
    assert "Phantom" in result.sheets
    assert any(c.ref == "A1" for c in result.cells)


def test_evaluate_formulas_stub_returns_false(tmp_path, xlsx_fixture):
    parser = ExcelDeterministicParser(tmp_dir=str(tmp_path / "out"))
    assert parser.evaluate_formulas(xlsx_fixture) is False


def test_zipfile_is_valid(tmp_path, xlsx_fixture):
    with zipfile.ZipFile(xlsx_fixture, "r") as zf:
        assert "xl/workbook.xml" in zf.namelist()

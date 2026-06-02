"""Generate two sample test-case spreadsheets for import testing.

Run from the backend/ directory:

    python samples/_generate_samples.py

Produces:
  - backend/samples/sample_test_cases_friendly_headers.xlsx
      AAOS-style human headers ("TC ID", "Reference Spec Document/s", ...)
      Comma-separated multi-values. Mixed coverage of all dropdown fields.

  - backend/samples/sample_test_cases_canonical_headers.xlsx
      snake_case canonical headers (matches DB columns 1:1).
      Pipe-separated multi-values. Includes minimal rows to exercise defaults.

Each workbook uses the sheet name "test_cases" so the importer auto-routes
to the test_cases target without needing an explicit selection.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

OUT_DIR = Path(__file__).resolve().parent

# -----------------------------------------------------------------------------
# File 1 — Friendly AAOS-style headers, comma-separated multi-values
# -----------------------------------------------------------------------------
FRIENDLY_HEADERS = [
    "TC ID",
    "Test Case Title",
    "Test Description",
    "Reference Spec Document/s",
    "Associated Requirement/s",
    "Screen ID/IDs",
    "Feature",
    "DR ID",
    "Preconditions",
    "Procedure",
    "Expected Behavior",
    "Test Type",
    "Region",
    "Brand",
    "Vehicle Variant",
    "Vehicle Mode",
    "Env Dependency",
    "Regulation",
    "Priority",
    "Test Suite Type",
]

FRIENDLY_ROWS = [
    [
        "TC_FOTA_0001",
        "FOTA — successful end-to-end OTA update",
        "Verify that a complete OTA update succeeds when the vehicle has good connectivity and the user accepts the update prompt.",
        "SPEC_FOTA_001, SPEC_OTA_FLOW_002",
        "REQ_FOTA_010, REQ_FOTA_011",
        "SCR_FOTA_HOME, SCR_FOTA_PROGRESS",
        "FOTA",
        "DR_FOTA_2024_07",
        "Vehicle ignition ON.\nWi-Fi connected.\nBattery > 50%.",
        "1. Open Settings > System > Updates.\n2. Tap 'Check for updates'.\n3. Accept the prompt.\n4. Wait for download + install.\n5. Reboot when prompted.",
        "Update completes successfully and the firmware version is bumped on the About screen.",
        "Positive",
        "US, CAN",
        "Ford, GM",
        "PZ1D, P33A",
        "Common, EV",
        "ECU, Vehicle",
        "Yes",
        "P1",
        "Sanity, Regression",
    ],
    [
        "TC_FOTA_0002",
        "FOTA — interrupted download recovers",
        "Confirm that an OTA download paused mid-transfer resumes correctly when connectivity is restored.",
        "SPEC_FOTA_001",
        "REQ_FOTA_012",
        "SCR_FOTA_PROGRESS",
        "FOTA",
        "",
        "OTA package partially downloaded (~50%).\nWi-Fi disconnected manually.",
        "1. Disconnect Wi-Fi during download.\n2. Wait 60s.\n3. Reconnect Wi-Fi.\n4. Confirm download resumes from the last byte.",
        "Download resumes and completes; no corruption is detected on install.",
        "Abnormal",
        "US, EU, JPN",
        "Tesla",
        "P33B, P33C",
        "EV, PHEV",
        "Simulator, HIL",
        "No",
        "P2",
        "Functional, Regression",
    ],
    [
        "TC_BT_0003",
        "Bluetooth — pair with Android phone",
        "Pair an Android device with the head unit and verify contacts/messages sync prompt is shown.",
        "SPEC_BT_PAIR_010",
        "REQ_BT_001",
        "SCR_BT_PAIR, SCR_BT_DEVICES",
        "BT",
        "DR_BT_2024_03",
        "Phone discoverable.\nHead unit BT enabled.",
        "1. Open BT settings.\n2. Tap 'Pair new device'.\n3. Select phone.\n4. Confirm passkey on both devices.",
        "Pairing succeeds; sync-prompt offers contacts and messages access.",
        "Positive",
        "US",
        "Honda, Toyota",
        "PZ1D",
        "ICE, HEV",
        "Vehicle",
        "No",
        "P2",
        "Sanity, Smoke",
    ],
    [
        "TC_RADIO_0004",
        "Radio — AM band scan with weak signal",
        "Verify the AM scan behavior in a low-signal area; head unit must skip stations below the SNR threshold.",
        "SPEC_RADIO_AM_004",
        "REQ_RADIO_021",
        "SCR_RADIO_AM",
        "Radio",
        "DR_RADIO_2024_11",
        "Vehicle parked in low-reception zone.",
        "1. Open Radio app.\n2. Switch to AM band.\n3. Trigger band scan.\n4. Observe stations added.",
        "Only stations with SNR ≥ threshold are added; UI shows 'No stations' if none qualify.",
        "Boundary",
        "MEX, CAN",
        "GM",
        "P33A",
        "Common",
        "Bench",
        "No",
        "P3",
        "Functional",
    ],
    [
        "TC_LAUNCHER_0005",
        "Launcher — pin a 3rd-party app",
        "Confirm that a side-loaded app can be pinned to the launcher home shortcut row.",
        "SPEC_LAUNCHER_PINS_002",
        "REQ_LAUNCHER_007, REQ_LAUNCHER_009",
        "SCR_LAUNCHER_HOME, SCR_LAUNCHER_ALL_APPS",
        "Launcher",
        "",
        "Side-loaded app installed via adb.",
        "1. Open All Apps.\n2. Long-press the side-loaded app.\n3. Tap 'Pin to home'.",
        "App icon appears on the home shortcut row and persists across reboots.",
        "Positive",
        "US, CAN, MEX",
        "Ford",
        "PZ2D",
        "Common, EV, HEV",
        "ECU",
        "No",
        "P3",
        "Functional, Regression",
    ],
    [
        "TC_NAV_0006",
        "Navigation — re-route on missed turn",
        "When the driver misses a turn, the navigation engine must recompute a route within 3 seconds.",
        "SPEC_NAV_REROUTE_001",
        "REQ_NAV_032",
        "SCR_NAV_MAP, SCR_NAV_DIRECTIONS",
        "Navigation",
        "DR_NAV_2024_05",
        "Active route in progress.",
        "1. Drive past a planned turn.\n2. Wait for the engine to detect off-route state.\n3. Observe new route appears.",
        "New route is rendered within 3 seconds and TBT directions update.",
        "Positive",
        "EU, JPN",
        "Tesla, Volvo",
        "P33B",
        "EV",
        "Vehicle, HIL",
        "Yes",
        "P1",
        "Regression, Acceptance",
    ],
    [
        "TC_CLIMATE_0007",
        "Climate — heat-pump fault notification",
        "When the heat-pump reports a hardware fault, the climate UI must display the fault banner and disable the heat-pump toggle.",
        "SPEC_CLIMATE_HP_003",
        "REQ_CLIMATE_018",
        "SCR_CLIMATE_HOME",
        "Climate",
        "DR_CLIMATE_2024_02",
        "Heat-pump simulator wired to fault state.",
        "1. Trigger fault state.\n2. Open Climate app.\n3. Inspect heat-pump toggle and banner.",
        "Banner reads 'Heat pump unavailable' and toggle is greyed-out.",
        "Negative",
        "US, EU",
        "GM, Ford",
        "PZ1D, PZ2D",
        "EV, PHEV",
        "Simulator",
        "No",
        "P2",
        "Functional",
    ],
    [
        "TC_PHONE_0008",
        "Phone — fall back to BT when LTE drops",
        "When the in-car eSIM loses LTE, outgoing calls must fall back to a paired phone over BT-HFP.",
        "SPEC_PHONE_FALLBACK_005",
        "REQ_PHONE_044",
        "SCR_PHONE_DIALER",
        "Phone",
        "",
        "Paired phone with BT-HFP active.\neSIM disabled.",
        "1. Place a call from dialer.\n2. Verify call is routed via paired phone.",
        "Call is placed and audio path uses BT-HFP; no fallback to local cellular.",
        "Abnormal",
        "US",
        "Honda",
        "P33A, P33C",
        "ICE",
        "ECU, Vehicle",
        "No",
        "P3",
        "Functional, Smoke",
    ],
    [
        "TC_MEDIA_0009",
        "Media — USB drive with mixed codecs",
        "Insert a USB drive containing FLAC, MP3, and AAC files; verify the playlist enumerates all of them.",
        "SPEC_MEDIA_USB_007",
        "REQ_MEDIA_022, REQ_MEDIA_023",
        "SCR_MEDIA_BROWSE, SCR_MEDIA_NOW_PLAYING",
        "Media",
        "DR_MEDIA_2024_09",
        "USB drive prepared with 3 codec types.",
        "1. Insert USB.\n2. Open Media > USB source.\n3. Browse playlist.",
        "All FLAC, MP3, and AAC tracks appear in the playlist.",
        "Positive",
        "APAC, JPN",
        "Toyota, Honda",
        "PZ1D, P33B",
        "HEV, ICE",
        "Bench, Vehicle",
        "No",
        "P3",
        "Sanity, Performance",
    ],
    [
        "TC_FOTA_0010",
        "FOTA — regulatory recall update is mandatory",
        "When the OTA package is flagged as a regulatory recall, the user cannot defer the install.",
        "SPEC_FOTA_RECALL_011",
        "REQ_FOTA_REG_001",
        "SCR_FOTA_PROGRESS, SCR_FOTA_RECALL_BANNER",
        "FOTA",
        "DR_FOTA_REG_2024_01",
        "Recall package staged for delivery.",
        "1. Receive recall package.\n2. Tap 'Defer'.\n3. Observe behavior.",
        "Defer button is disabled; UI explains the install is mandatory by regulation.",
        "Negative",
        "US, CAN, MEX, EU",
        "Ford, GM, Tesla, Toyota, Honda, Volvo",
        "PZ1D, PZ2D, P33A, P33B, P33C",
        "Common, EV, HEV, ICE, PHEV",
        "ECU, Simulator, HIL, Bench, Vehicle",
        "Yes",
        "P1",
        "Sanity, Functional, Regression, Acceptance",
    ],
]


# -----------------------------------------------------------------------------
# File 2 — Canonical snake_case headers, pipe-separated multi-values
# -----------------------------------------------------------------------------
CANONICAL_HEADERS = [
    "test_case_id",
    "title",
    "description",
    "reference_document",
    "associated_requirement_id",
    "screen_id",
    "feature",
    "dr_id",
    "test_objective",
    "preconditions",
    "procedure",
    "expected_behavior",
    "test_type",
    "region",
    "brand",
    "vehicle_variant",
    "vehicle_mode",
    "env_dependency",
    "regulation",
    "priority",
    "testsuite_type",
]

CANONICAL_ROWS = [
    [
        "TC_BT_0101",
        "BT — pair iPhone with auto-accept",
        "iPhone pairing must auto-accept the passkey when the vehicle is in dealer-mode.",
        "SPEC_BT_PAIR_010|SPEC_BT_DEALER_004",
        "REQ_BT_002|REQ_BT_DEALER_001",
        "SCR_BT_PAIR",
        "BT",
        "DR_BT_2024_04",
        "Validate dealer-mode auto-accept pairing for iPhones.",
        "Vehicle in dealer-mode.\niPhone discoverable.",
        "1. Initiate pairing.\n2. Confirm head-unit auto-accepts.",
        "Pairing succeeds without user prompt on either side.",
        "Positive",
        "US|CAN",
        "Ford|GM",
        "PZ1D|PZ2D",
        "EV|PHEV",
        "Vehicle",
        "No",
        "P2",
        "Sanity|Functional",
    ],
    [
        "TC_RADIO_0102",
        "Radio — HD seek skips analog-only stations",
        "",  # description intentionally blank — exercise default + minimal row paths
        "SPEC_RADIO_HD_006",
        "REQ_RADIO_028",
        "SCR_RADIO_FM",
        "Radio",
        "",
        "",
        "FM band, HD-mode toggle ON.",
        "1. Trigger seek.\n2. Observe behavior.",
        "Seek skips analog-only stations; only HD-capable channels are landed on.",
        "Boundary",
        "US",
        "Tesla",
        "P33A",
        "EV",
        "Bench",
        "No",
        "P3",
        "Functional",
    ],
    [
        "TC_LAUNCHER_0103",
        "Launcher — locale switch updates app labels",
        "Switching the system locale must trigger an immediate re-render of all app labels on the launcher.",
        "SPEC_LAUNCHER_I18N_001",
        "REQ_LAUNCHER_010|REQ_I18N_001",
        "SCR_LAUNCHER_HOME|SCR_SETTINGS_LANG",
        "Launcher",
        "DR_LAUNCHER_I18N_2024_01",
        "",
        "Settings open at the language picker.",
        "1. Pick Japanese.\n2. Return to launcher.\n3. Confirm labels are localized.",
        "All system app labels render in Japanese without a reboot.",
        "Positive",
        "JPN|APAC",
        "Toyota|Honda",
        "P33B|P33C",
        "ICE|HEV",
        "Vehicle",
        "No",
        "P2",
        "Functional|Regression",
    ],
    [
        "TC_NAV_0104",
        "Navigation — POI search with diacritics",
        "Search for a POI containing accented characters (e.g. 'Café Düsseldorf') and confirm the result list is non-empty.",
        "SPEC_NAV_SEARCH_002",
        "REQ_NAV_039",
        "SCR_NAV_SEARCH",
        "Navigation",
        "",
        "",
        "Map data: EU.",
        "1. Open POI search.\n2. Enter 'Café Düsseldorf'.\n3. Inspect results.",
        "Result list contains the matching POI.",
        "Positive",
        "EU",
        "Volvo",
        "P33C",
        "EV|HEV",
        "Simulator",
        "No",
        "P3",
        "Functional",
    ],
    [
        "TC_CLIMATE_0105",
        "Climate — boundary at max defrost",
        "When defrost is at max for ≥10 minutes, the climate engine must throttle to protect the heater core.",
        "SPEC_CLIMATE_DEFROST_004",
        "REQ_CLIMATE_021",
        "SCR_CLIMATE_HOME",
        "Climate",
        "DR_CLIMATE_2024_03",
        "Validate defrost protection logic.",
        "Cabin temp simulator at -10°C.",
        "1. Set defrost to MAX.\n2. Wait 10 minutes.\n3. Inspect heater output.",
        "Heater output is throttled to protective level; UI shows discrete indicator.",
        "Boundary",
        "CAN|JPN",
        "Toyota|Honda|Volvo",
        "P33A|P33B",
        "Common",
        "HIL|Bench",
        "No",
        "P2",
        "Functional|Regression",
    ],
    [
        "TC_PHONE_0106",
        "Phone — emergency call from locked head unit",
        "Verify that the dialer can place an emergency call even when the head unit is locked.",
        "SPEC_PHONE_E911_001",
        "REQ_PHONE_E911",
        "SCR_PHONE_DIALER|SCR_LOCK",
        "Phone",
        "DR_PHONE_E911_2024_01",
        "Emergency-call regulatory requirement.",
        "Head unit locked.\nValid eSIM.",
        "1. Tap emergency icon on lock screen.\n2. Place call to test E911 endpoint.",
        "Call is placed without unlocking; correct location data is included.",
        "Positive",
        "US|CAN|MEX",
        "Ford|GM|Tesla",
        "PZ1D|PZ2D|P33A|P33B|P33C",
        "Common|EV|HEV|ICE|PHEV",
        "Vehicle",
        "Yes",
        "P1",
        "Sanity|Acceptance",
    ],
    [
        "TC_MEDIA_0107",
        "Media — gapless playback FLAC",
        "Play a FLAC album marked 'gapless' and verify there is no audible silence between tracks.",
        "SPEC_MEDIA_GAPLESS_002",
        "REQ_MEDIA_030",
        "SCR_MEDIA_NOW_PLAYING",
        "Media",
        "",
        "",
        "USB with gapless album.",
        "1. Play album.\n2. Listen at track boundaries.",
        "No audible silence detected at boundaries.",
        "Positive",
        "EU|APAC",
        "Tesla|Volvo",
        "PZ2D|P33C",
        "EV",
        "Bench",
        "No",
        "P4",
        "Performance",
    ],
    [
        "TC_FOTA_0108",
        "FOTA — abnormal: package signature mismatch",
        "If the OTA package signature does not match the trusted root, the install must be aborted before flashing.",
        "SPEC_FOTA_SEC_002|SPEC_FOTA_RECALL_011",
        "REQ_FOTA_SEC_001",
        "SCR_FOTA_PROGRESS",
        "FOTA",
        "DR_FOTA_SEC_2024_02",
        "Negative-path security validation.",
        "Tampered OTA package staged.",
        "1. Trigger install.\n2. Observe abort path.",
        "Install aborts with security error; firmware version unchanged.",
        "Abnormal",
        "US|EU",
        "Ford|GM|Tesla",
        "PZ1D|P33A|P33B",
        "Common|EV",
        "Simulator|HIL",
        "Yes",
        "P1",
        "Functional|Regression|Acceptance",
    ],
    [
        "TC_MIN_0109",
        "Minimal row — defaults applied",
        "",  # exercise priority/test_type defaults from TARGET_CONFIG
        "",
        "",
        "",
        "Launcher",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ],
    [
        "TC_FOTA_0001",  # collides with friendly file → exercise upsert/duplicate path
        "FOTA — successful end-to-end OTA update (canonical edit)",
        "Updated description for the same TC ID; importer should treat this as an update or report a duplicate cleanly.",
        "SPEC_FOTA_001",
        "REQ_FOTA_010",
        "SCR_FOTA_HOME",
        "FOTA",
        "DR_FOTA_2024_07",
        "Re-import same TC to test upsert.",
        "Same as original.",
        "Same as original.",
        "Updated expected behavior text.",
        "Positive",
        "US",
        "Ford",
        "PZ1D",
        "Common",
        "Vehicle",
        "Yes",
        "P1",
        "Regression",
    ],
]


# -----------------------------------------------------------------------------
# Common workbook helpers
# -----------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(bold=True)


def _write_workbook(path: Path, headers: list, rows: list) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "test_cases"

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    for col_idx, header in enumerate(headers, start=1):
        # Crude width heuristic: longest cell in the column, capped, padded.
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            [len(str(header))]
            + [len(str(r[col_idx - 1])) if col_idx - 1 < len(r) else 0 for r in rows]
        )
        ws.column_dimensions[col_letter].width = min(max(max_len, 12), 60)

    # Wrap-text for the body rows so multi-line procedures render readable.
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    friendly = OUT_DIR / "sample_test_cases_friendly_headers.xlsx"
    canonical = OUT_DIR / "sample_test_cases_canonical_headers.xlsx"

    _write_workbook(friendly, FRIENDLY_HEADERS, FRIENDLY_ROWS)
    print(f"wrote {friendly}  rows={len(FRIENDLY_ROWS)}")

    _write_workbook(canonical, CANONICAL_HEADERS, CANONICAL_ROWS)
    print(f"wrote {canonical}  rows={len(CANONICAL_ROWS)}")


if __name__ == "__main__":
    main()

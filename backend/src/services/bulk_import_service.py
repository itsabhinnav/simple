import csv
from dataclasses import dataclass
from pathlib import Path
import re
import tempfile
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

# File extensions we accept for bulk import. Extend here, then teach
# `_parse_file_to_sheets` how to read the new format.
SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".csv"}

from src.infrastructure.dependency_injection import get_hybrid_database_service
from src.infrastructure.logging_config import get_logger

logger = get_logger(__name__)


TARGET_ALIASES = {
    "spec": "specifications",
    "specs": "specifications",
    "specification": "specifications",
    "specifications": "specifications",
    "requirement": "requirements",
    "requirements": "requirements",
    "req": "requirements",
    "reqs": "requirements",
    "test": "test_cases",
    "tests": "test_cases",
    "testcase": "test_cases",
    "testcases": "test_cases",
    "test_case": "test_cases",
    "test_cases": "test_cases",
    "design": "design_tickets",
    "designs": "design_tickets",
    "design_ticket": "design_tickets",
    "design_tickets": "design_tickets",
    "ticket": "design_tickets",
    "tickets": "design_tickets",
}


HEADER_ALIASES = {
    "id": None,
    "specification_id": "spec_id",
    "spec id": "spec_id",
    "requirement id": "requirement_id",
    "req id": "requirement_id",
    "test case id": "test_case_id",
    "testcase id": "test_case_id",
    "tc id": "test_case_id",
    "tc_id": "test_case_id",
    "test id": "test_case_id",
    "design ticket id": "design_ticket_id",
    "design id": "design_ticket_id",
    "when": "when_action",
    "then": "then_result",
    "expected result": "then_result",
    "expected": "expected_behavior",
    "expected behavior": "expected_behavior",
    "expected behaviour": "expected_behavior",
    "expected output": "expected_behavior",
    "expected outcome": "expected_behavior",
    "linked requirement": "linked_requirement_id",
    "linked requirement id": "linked_requirement_id",
    "associated requirement": "associated_requirement_id",
    "associated requirement id": "associated_requirement_id",
    "image": "image_url",
    "file": "file_url",
    # ---- Aliases for the test-case import (non-standard AAOS spreadsheets) ----
    "test title": "title",
    "test case title": "title",
    "testcase title": "title",
    "title": "title",
    "summary": "title",
    "name": "title",
    "test name": "title",
    "description": "description",
    "test description": "description",
    "testcase description": "description",
    "details": "description",
    "vehicle": "vehicle_model",
    "vehicle model": "vehicle_model",
    "model": "vehicle_model",
    "car model": "vehicle_model",
    "platform": "vehicle_model",
    "severity": "severity",
    "criticality": "severity",
    "impact": "severity",
    "test severity": "severity",
    "objective": "test_objective",
    "test objective": "test_objective",
    "preconditions": "preconditions",
    "pre conditions": "preconditions",
    "pre-conditions": "preconditions",
    "pre requisites": "preconditions",
    "prerequisites": "preconditions",
    "setup": "preconditions",
    "steps": "procedure",
    "test steps": "procedure",
    "procedure": "procedure",
    "test procedure": "procedure",
    "feature": "feature",
    "feature name": "feature",
    "aaos feature": "feature",
    "module": "feature",
    "screen id": "screen_id",
    "screen ids": "screen_id",
    "screen id/ids": "screen_id",
    "screen id ids": "screen_id",
    "screen id/s": "screen_id",
    "screens": "screen_id",
    "screen": "screen_id",
    "priority": "priority",
    "test priority": "priority",
    "test type": "test_type",
    "type": "test_type",
    "region": "region",
    "regions": "region",
    "brand": "brand",
    "brands": "brand",
    "oem": "brand",
    "oems": "brand",
    "vehicle variant": "vehicle_variant",
    "vehicle variants": "vehicle_variant",
    "variant": "vehicle_variant",
    "variants": "vehicle_variant",
    # vehicle_mode is the canonical multi-value powertrain field
    # (Common/EV/HEV/ICE/PHEV). vehicle_specification is a separate free-form
    # legacy column kept for backwards compatibility — do NOT alias the same
    # human header to both, or one mapping silently shadows the other.
    "vehicle mode": "vehicle_mode",
    "vehicle modes": "vehicle_mode",
    "vehiclemode": "vehicle_mode",
    "powertrain": "vehicle_mode",
    "drive mode": "vehicle_mode",
    "vehicle specification": "vehicle_specification",
    "vehicle spec": "vehicle_specification",
    "env dependency": "env_dependency",
    "env dependencies": "env_dependency",
    "environment": "env_dependency",
    "environment dependency": "env_dependency",
    "environment dependencies": "env_dependency",
    "regulation": "regulation",
    "regulations": "regulation",
    "is regulation": "regulation",
    "is regulation test case": "regulation",
    "regulation test case": "regulation",
    "compliance": "regulation",
    "testsuite type": "testsuite_type",
    "test suite type": "testsuite_type",
    "test suite": "testsuite_type",
    "test suites": "testsuite_type",
    "suite": "testsuite_type",
    "reference document": "reference_document",
    "reference documents": "reference_document",
    "reference document/s": "reference_document",
    "reference doc": "reference_document",
    "reference docs": "reference_document",
    "reference": "reference_document",
    "ref doc": "reference_document",
    "ref docs": "reference_document",
    "ref document": "reference_document",
    "spec document": "reference_document",
    "spec documents": "reference_document",
    "spec doc": "reference_document",
    "spec docs": "reference_document",
    "reference spec": "reference_document",
    "reference spec doc": "reference_document",
    "reference spec docs": "reference_document",
    "reference spec document": "reference_document",
    "reference spec documents": "reference_document",
    "reference spec document/s": "reference_document",
    "associated requirement/s": "associated_requirement_id",
    "associated requirements": "associated_requirement_id",
    "associated requirement ids": "associated_requirement_id",
    "associated requirement id/s": "associated_requirement_id",
    "linked requirements": "linked_requirement_id",
    "dr id": "dr_id",
    "drid": "dr_id",
    "design review id": "dr_id",
    "design review": "dr_id",
    "dr applicable screens": "dr_applicable_screens",
    "dr screens": "dr_applicable_screens",
}


TARGET_CONFIG = {
    "specifications": {
        "table": "specifications",
        "id_field": "spec_id",
        "prefix": "SPEC",
        "required": ["spec_id", "title"],
        "defaults": {"status": "Draft"},
        "fields": ["spec_id", "title", "description", "category", "version", "status", "file_url", "created_by"],
        "ddl": """
            CREATE TABLE IF NOT EXISTS specifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                spec_id TEXT UNIQUE,
                title TEXT NOT NULL,
                description TEXT,
                category TEXT,
                version TEXT,
                status TEXT,
                file_url TEXT,
                created_by TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """,
    },
    "requirements": {
        "table": "requirements",
        "id_field": "requirement_id",
        "prefix": "REQ",
        "required": ["requirement_id", "title"],
        "defaults": {"priority": "P2", "status": "Draft"},
        "fields": [
            "requirement_id", "title", "description", "requirement_type", "given",
            "when_action", "then_result", "priority", "status", "assignee", "tags", "created_by",
        ],
        "ddl": None,
    },
    "test_cases": {
        "table": "test_cases",
        "id_field": "test_case_id",
        "prefix": "TC",
        "required": ["test_case_id"],
        "defaults": {"priority": "P2", "test_type": "Positive"},
        "fields": [
            "test_case_id", "title", "description", "vehicle_model", "severity",
            "reference_document", "associated_requirement_id", "screen_id",
            "feature", "dr_applicable_screens", "dr_id", "test_objective", "preconditions",
            "procedure", "expected_behavior", "test_type", "region", "brand", "vehicle_variant",
            "vehicle_specification", "vehicle_mode", "env_dependency", "requirement_type",
            "regulation", "priority", "testsuite_type", "created_by",
        ],
        "ddl": None,
    },
    "design_tickets": {
        "table": "design_tickets",
        "id_field": "design_ticket_id",
        "prefix": "DT",
        "required": ["design_ticket_id", "title"],
        "defaults": {"priority": "P2", "status": "Draft"},
        "fields": [
            "design_ticket_id", "title", "description", "design_type", "diagram_type",
            "image_url", "priority", "status", "linked_requirement_id", "assignee", "tags", "created_by",
        ],
        "ddl": """
            CREATE TABLE IF NOT EXISTS design_tickets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                design_ticket_id TEXT UNIQUE NOT NULL,
                title TEXT NOT NULL,
                description TEXT,
                design_type TEXT,
                diagram_type TEXT,
                image_url TEXT,
                priority TEXT,
                status TEXT DEFAULT 'Draft',
                linked_requirement_id TEXT,
                assignee TEXT,
                tags TEXT,
                created_by TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """,
    },
}


@dataclass
class ImportRowResult:
    row: int
    status: str
    id: Optional[str] = None
    error: Optional[str] = None


class _CsvSheet:
    """Quack-alike of openpyxl's Worksheet, just enough to feed into the
    existing pipeline. Used so CSV imports reuse the same row/header
    machinery as xlsx imports without bifurcating the code paths.
    """

    def __init__(self, title: str, rows: List[List[Any]]):
        self.title = title
        self._rows = rows
        # `max_row` mirrors openpyxl's 1-indexed total-rows attribute; the
        # preview path subtracts 1 to estimate body row count.
        self.max_row = len(rows)

    def iter_rows(self, values_only: bool = True):
        for row in self._rows:
            yield tuple(row)


class BulkImportService:
    def __init__(self):
        self.database_service = get_hybrid_database_service()

    # ------------------------------------------------------------------
    # Preview / mapping APIs
    # ------------------------------------------------------------------
    def get_target_fields(self, target: str) -> Dict[str, Any]:
        """Return the known canonical fields for a given target so the
        frontend can render a mapping UI."""
        normalized = self._normalize_target(target)
        if normalized is None:
            raise ValueError("Unsupported import type")
        config = TARGET_CONFIG[normalized]
        return {
            "target": normalized,
            "id_field": config["id_field"],
            "required": config["required"],
            "fields": config["fields"],
        }

    def preview_file(self, file_storage, target: str, sample_rows: int = 5) -> Dict[str, Any]:
        """Parse the uploaded workbook and return a per-sheet preview with
        detected headers, auto-detected mapping, and a few sample rows. The
        file is NOT imported. The frontend uses this to render a mapping UI
        for non-standard test-case spreadsheets (10k+ rows, many vendors).

        Accepts ``.xlsx``/``.xlsm`` workbooks (multi-sheet) and ``.csv``
        files (single sheet). CSV files derive their "sheet name" from the
        filename stem so the auto-target inference still works.
        """
        normalized_target = self._normalize_target(target)
        if normalized_target is None and target != "auto":
            raise ValueError("Unsupported import type")

        filename = file_storage.filename or ""
        suffix = Path(filename).suffix.lower()
        if suffix not in SUPPORTED_SUFFIXES:
            raise ValueError(
                f"Unsupported file type '{suffix}'. Supported: "
                f"{', '.join(sorted(SUPPORTED_SUFFIXES))}"
            )

        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            file_storage.save(tmp.name)
            temp_path = tmp.name

        try:
            sheets = self._open_sheets(temp_path, filename)
            sheets_preview = []
            for sheet in sheets:
                sheet_target = normalized_target or self._infer_target(sheet.title, filename)
                config = TARGET_CONFIG.get(sheet_target) if sheet_target else None
                rows = sheet.iter_rows(values_only=True)
                header_row = next(rows, None) or []
                raw_headers = [str(h).strip() if h is not None else "" for h in header_row]
                detected = self._normalize_headers(header_row)
                suggested_mapping = {
                    raw: detected[i] if i < len(detected) else None
                    for i, raw in enumerate(raw_headers)
                    if raw
                }
                samples = []
                for idx, row_values in enumerate(rows):
                    if idx >= sample_rows:
                        break
                    samples.append([self._clean_cell(v) for v in row_values])
                sheets_preview.append({
                    "sheet": sheet.title,
                    "target": sheet_target,
                    "row_count_estimate": sheet.max_row - 1 if sheet.max_row else 0,
                    "raw_headers": raw_headers,
                    "suggested_mapping": suggested_mapping,
                    "known_fields": config["fields"] if config else [],
                    "id_field": config["id_field"] if config else None,
                    "required": config["required"] if config else [],
                    "sample_rows": samples,
                })
            return {"file": filename, "sheets": sheets_preview}
        finally:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass

    def import_files_with_mapping(
        self,
        files,
        target: str,
        mapping: Dict[str, str],
        created_by: str = "system",
        duplicate_strategy: str = "skip",
    ) -> Dict[str, Any]:
        """Import a workbook using a user-provided header→field mapping.

        `mapping` keys are the raw header strings exactly as they appear in
        the spreadsheet; values are the canonical field names from
        `TARGET_CONFIG[target]["fields"]`. Headers absent from the mapping
        fall back to auto-detection so users only need to override the
        ambiguous columns.

        `duplicate_strategy` controls what happens for rows whose primary
        ID is already present:

        - ``"skip"``    (default) — leave the existing row alone and count
          it under ``skipped``.
        - ``"replace"`` — UPDATE the existing row with the spreadsheet
          values and count it under ``updated``.
        """
        normalized_target = self._normalize_target(target)
        if normalized_target is None:
            raise ValueError("Unsupported import type")
        strategy = self._normalize_strategy(duplicate_strategy)
        normalized_mapping = {
            self._normalize_header(str(raw)): value
            for raw, value in (mapping or {}).items()
            if value
        }
        summaries = []
        totals = {"created": 0, "updated": 0, "skipped": 0, "failed": 0}
        for file_storage in files:
            filename = file_storage.filename or ""
            suffix = Path(filename).suffix.lower()
            if suffix not in SUPPORTED_SUFFIXES:
                summaries.append({
                    "file": filename, "created": 0, "skipped": 0, "failed": 1,
                    "sheets": [],
                    "errors": [{
                        "row": None,
                        "error": f"Unsupported file type '{suffix}'. Supported: "
                                 f"{', '.join(sorted(SUPPORTED_SUFFIXES))}",
                    }],
                })
                totals["failed"] += 1
                continue
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                file_storage.save(tmp.name)
                temp_path = tmp.name
            try:
                summary = self._import_workbook(
                    temp_path, filename, normalized_target, created_by,
                    user_mapping=normalized_mapping,
                    duplicate_strategy=strategy,
                )
                summaries.append(summary)
                for key in totals:
                    totals[key] += summary[key]
            finally:
                try:
                    Path(temp_path).unlink(missing_ok=True)
                except Exception:
                    pass
        return {"files": summaries, "totals": totals}

    def import_files(
        self,
        files,
        target: str,
        created_by: str = "system",
        duplicate_strategy: str = "skip",
    ) -> Dict[str, Any]:
        normalized_target = self._normalize_target(target)
        if normalized_target is None and target != "auto":
            raise ValueError("Unsupported import type")
        strategy = self._normalize_strategy(duplicate_strategy)

        summaries = []
        totals = {"created": 0, "updated": 0, "skipped": 0, "failed": 0}

        for file_storage in files:
            filename = file_storage.filename or ""
            suffix = Path(filename).suffix.lower()
            if suffix not in SUPPORTED_SUFFIXES:
                summaries.append({
                    "file": filename,
                    "created": 0,
                    "skipped": 0,
                    "failed": 1,
                    "sheets": [],
                    "errors": [{
                        "row": None,
                        "error": f"Unsupported file type '{suffix}'. Supported: "
                                 f"{', '.join(sorted(SUPPORTED_SUFFIXES))}",
                    }],
                })
                totals["failed"] += 1
                continue

            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                file_storage.save(tmp.name)
                temp_path = tmp.name

            try:
                summary = self._import_workbook(
                    temp_path, filename, normalized_target, created_by,
                    duplicate_strategy=strategy,
                )
                summaries.append(summary)
                for key in totals:
                    totals[key] += summary[key]
            finally:
                try:
                    Path(temp_path).unlink(missing_ok=True)
                except Exception:
                    pass

        return {"files": summaries, "totals": totals}

    def _import_workbook(
        self,
        path: str,
        filename: str,
        target: Optional[str],
        created_by: str,
        user_mapping: Optional[Dict[str, str]] = None,
        duplicate_strategy: str = "skip",
    ) -> Dict[str, Any]:
        sheets = self._open_sheets(path, filename)
        file_summary = {
            "file": filename, "created": 0, "updated": 0, "skipped": 0, "failed": 0,
            "sheets": [], "errors": [],
        }

        for sheet in sheets:
            sheet_target = target or self._infer_target(sheet.title, filename)
            if not sheet_target:
                continue

            sheet_summary = self._import_sheet(
                sheet, sheet_target, created_by,
                user_mapping=user_mapping,
                duplicate_strategy=duplicate_strategy,
            )
            file_summary["sheets"].append(sheet_summary)
            file_summary["created"] += sheet_summary["created"]
            file_summary["updated"] += sheet_summary["updated"]
            file_summary["skipped"] += sheet_summary["skipped"]
            file_summary["failed"] += sheet_summary["failed"]
            file_summary["errors"].extend([
                {"sheet": sheet.title, **error} for error in sheet_summary["errors"][:20]
            ])

        if not file_summary["sheets"]:
            file_summary["failed"] += 1
            file_summary["errors"].append({"row": None, "error": "No matching sheets found for the selected import type"})

        return file_summary

    def _import_sheet(
        self,
        sheet,
        target: str,
        created_by: str,
        user_mapping: Optional[Dict[str, str]] = None,
        duplicate_strategy: str = "skip",
    ) -> Dict[str, Any]:
        config = TARGET_CONFIG[target]
        self._ensure_table(config)

        rows = sheet.iter_rows(values_only=True)
        header_row = next(rows, None)
        headers = self._normalize_headers(header_row or [], user_mapping=user_mapping)
        summary = {
            "sheet": sheet.title, "target": target,
            "created": 0, "updated": 0, "skipped": 0, "failed": 0, "errors": [],
        }

        if not any(headers):
            summary["failed"] += 1
            summary["errors"].append({"row": 1, "error": "Header row is empty"})
            return summary

        for row_number, row_values in enumerate(rows, start=2):
            if self._is_blank_row(row_values):
                continue
            row_data = self._row_to_dict(headers, row_values)
            result = self._import_row(config, row_data, row_number, created_by, duplicate_strategy)
            summary[result.status] += 1
            if result.error:
                summary["errors"].append({"row": result.row, "id": result.id, "error": result.error})

        return summary

    def _import_row(
        self,
        config: Dict[str, Any],
        row_data: Dict[str, Any],
        row_number: int,
        created_by: str,
        duplicate_strategy: str = "skip",
    ) -> ImportRowResult:
        try:
            payload = self._prepare_payload(config, row_data, row_number, created_by)
            missing = [field for field in config["required"] if not payload.get(field)]
            if missing:
                return ImportRowResult(row=row_number, status="failed", error=f"Missing required fields: {', '.join(missing)}")

            table = config["table"]
            id_field = config["id_field"]
            record_id = payload[id_field]
            existing = self.database_service.execute_query(
                f"SELECT id FROM {table} WHERE {id_field} = ?",
                "default",
                params=(record_id,),
            )
            if existing.get("data"):
                if duplicate_strategy == "replace":
                    # UPDATE only the non-None payload fields so empty cells in
                    # the spreadsheet don't blank out previously-saved values
                    # on the existing row. id_field stays out of the SET list.
                    update_fields = [
                        field for field in config["fields"]
                        if field != id_field and payload.get(field) is not None
                    ]
                    if not update_fields:
                        return ImportRowResult(
                            row=row_number, status="skipped", id=record_id,
                            error="Row had no non-empty fields to update",
                        )
                    set_clause = ", ".join(f"{f} = ?" for f in update_fields)
                    values = tuple(payload[f] for f in update_fields) + (record_id,)
                    upd = self.database_service.execute_query(
                        f"UPDATE {table} SET {set_clause}, updated_at = CURRENT_TIMESTAMP "
                        f"WHERE {id_field} = ?",
                        "default",
                        params=values,
                    )
                    if not upd.get("success"):
                        return ImportRowResult(
                            row=row_number, status="failed", id=record_id,
                            error=upd.get("error", "Update failed"),
                        )
                    return ImportRowResult(row=row_number, status="updated", id=record_id)
                return ImportRowResult(row=row_number, status="skipped", id=record_id, error="Duplicate ID skipped")

            fields = [field for field in config["fields"] if payload.get(field) is not None]
            placeholders = ", ".join(["?" for _ in fields])
            columns = ", ".join(fields)
            values = tuple(payload[field] for field in fields)
            result = self.database_service.execute_query(
                f"INSERT INTO {table} ({columns}, created_at, updated_at) VALUES ({placeholders}, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)",
                "default",
                params=values,
            )
            if not result.get("success"):
                return ImportRowResult(row=row_number, status="failed", id=record_id, error=result.get("error", "Insert failed"))

            return ImportRowResult(row=row_number, status="created", id=record_id)
        except Exception as exc:
            logger.exception("Bulk import row failed")
            return ImportRowResult(row=row_number, status="failed", error=str(exc))

    @staticmethod
    def _normalize_strategy(value: Optional[str]) -> str:
        """Accept ``skip``/``replace`` (case-insensitive) and a few common
        aliases. Anything else falls back to the safe default ``skip``."""
        if not value:
            return "skip"
        v = str(value).strip().lower()
        if v in {"replace", "overwrite", "update", "upsert"}:
            return "replace"
        return "skip"

    def _prepare_payload(self, config: Dict[str, Any], row_data: Dict[str, Any], row_number: int, created_by: str) -> Dict[str, Any]:
        payload = {field: row_data.get(field) for field in config["fields"]}
        payload.update({key: value for key, value in config["defaults"].items() if not payload.get(key)})
        payload["created_by"] = row_data.get("created_by") or created_by

        id_field = config["id_field"]
        if not payload.get(id_field):
            payload[id_field] = self._generated_id(config["prefix"], row_number)

        if "title" in config["required"] and not payload.get("title"):
            payload["title"] = payload.get("description") or payload.get(id_field)

        return payload

    def _open_sheets(self, path: str, filename: str):
        """Open ``path`` and return an iterable of sheet-like objects that
        the import/preview pipeline can consume uniformly. Supports xlsx,
        xlsm, and csv.

        For CSV we synthesise a single "sheet" named after the file stem
        (e.g. ``aaos_tests`` for ``aaos_tests.csv``) so the auto-target
        inference from `_infer_target` still works the same way.
        """
        suffix = Path(filename).suffix.lower()
        if suffix == ".csv":
            return [self._read_csv_as_sheet(path, filename)]
        # Default: openpyxl-readable workbook with one or more sheets.
        workbook = load_workbook(path, read_only=True, data_only=True)
        return list(workbook.worksheets)

    def _read_csv_as_sheet(self, path: str, filename: str) -> "_CsvSheet":
        """Read a CSV file into the `_CsvSheet` shim. Sniffs the dialect
        (delimiter, quoting) and falls back to standard CSV defaults if
        sniffing fails. Tries UTF-8 with BOM fallback so files exported
        from Excel still parse cleanly."""
        rows: List[List[Any]] = []
        # Read a small chunk first so we can sniff the dialect without
        # holding the whole file in memory for very large CSVs.
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                with open(path, "r", encoding=encoding, newline="") as f:
                    sample = f.read(4096)
                    f.seek(0)
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                    except csv.Error:
                        dialect = csv.excel
                    reader = csv.reader(f, dialect)
                    for row in reader:
                        rows.append(list(row))
                break
            except UnicodeDecodeError:
                rows = []
                continue
        else:
            raise ValueError(f"Could not decode CSV file '{filename}' with any supported encoding")
        title = Path(filename).stem or "csv"
        return _CsvSheet(title=title, rows=rows)

    def _ensure_table(self, config: Dict[str, Any]) -> None:
        ddl = config.get("ddl")
        if ddl:
            self.database_service.execute_query(ddl, "default")

    def _normalize_target(self, target: str) -> Optional[str]:
        key = self._normalize_header(target or "")
        return TARGET_ALIASES.get(key)

    def _infer_target(self, sheet_name: str, filename: str) -> Optional[str]:
        for value in (sheet_name, Path(filename).stem):
            normalized = self._normalize_header(value)
            for alias, target in TARGET_ALIASES.items():
                if alias in normalized:
                    return target
        return None

    def _normalize_headers(
        self,
        headers: List[Any],
        user_mapping: Optional[Dict[str, str]] = None,
    ) -> List[Optional[str]]:
        return [self._normalize_header_cell(header, user_mapping) for header in headers]

    def _normalize_header_cell(
        self,
        header: Any,
        user_mapping: Optional[Dict[str, str]] = None,
    ) -> Optional[str]:
        if header is None:
            return None
        raw = str(header).strip()
        if not raw:
            return None
        # Explicit user-provided mapping wins over aliases / heuristics so the
        # frontend mapping UI is authoritative.
        if user_mapping:
            normalized_raw = self._normalize_header(raw)
            if normalized_raw in user_mapping:
                mapped = user_mapping[normalized_raw]
                return mapped or None
        lowered = raw.lower().strip()
        if lowered in HEADER_ALIASES:
            return HEADER_ALIASES[lowered]
        return self._normalize_header(raw)

    def _normalize_header(self, value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")

    def _row_to_dict(self, headers: List[Optional[str]], row_values: List[Any]) -> Dict[str, Any]:
        row = {}
        for header, value in zip(headers, row_values):
            if not header:
                continue
            clean_value = self._clean_cell(value)
            if clean_value is not None:
                row[header] = clean_value
        return row

    def _clean_cell(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            return value or None
        return str(value).strip()

    def _is_blank_row(self, row_values: List[Any]) -> bool:
        return all(self._clean_cell(value) is None for value in row_values)

    def _generated_id(self, prefix: str, row_number: int) -> str:
        import time
        return f"{prefix}_{int(time.time())}_{row_number}"
